"""Commission payouts: what each agent has actually been paid against what
they earned on a deal.

deal_agents (see db.py) already records who earned what on every deal —
views_deals.py writes it when a deal is saved. This module is the
accountant's layer on top of that same data: the payment records, and the
running paid/outstanding figures built from them. Nothing here duplicates
a deal's own commission math; it only tracks money actually handed over.
"""
from openpyxl import load_workbook

from db import execute, log, now, query, to_utc


def _status(earned, paid):
    if paid <= 0:
        return "unpaid"
    if paid + 0.005 >= earned:
        return "paid"
    return "partial"


def payouts(status="", agent_id=None, q=""):
    """Every agent's commission share on a live deal, with what's been paid
    against it. Cancelled deals carry no payable commission, so they're
    left out entirely rather than showing as a permanently 'unpaid' row."""
    where, args = ["d.status != 'Cancelled'"], []
    if status:
        where.append("da.payout_status = ?")
        args.append(status)
    if agent_id:
        where.append("da.user_id = ?")
        args.append(agent_id)
    if q:
        where.append("(d.ref LIKE ? OR u.name LIKE ? OR p.title LIKE ?)")
        args += [f"%{q}%"] * 3
    sql = (
        "SELECT da.*, d.ref, d.status AS deal_status, d.closed_at, d.value,"
        " d.deal_type, u.name AS agent_name, p.title AS prop_title"
        " FROM deal_agents da"
        " JOIN deals d ON d.id = da.deal_id"
        " JOIN users u ON u.id = da.user_id"
        " LEFT JOIN properties p ON p.id = d.property_id"
        " WHERE " + " AND ".join(where) +
        " ORDER BY d.closed_at DESC, d.id DESC")
    return query(sql, args)


def summary(rows):
    earned = sum(r["amount"] or 0 for r in rows)
    paid = sum(r["paid_amount"] or 0 for r in rows)
    return {
        "earned": earned,
        "paid": paid,
        "outstanding": earned - paid,
        "unpaid": sum(1 for r in rows if r["payout_status"] == "unpaid"),
        "partial": sum(1 for r in rows if r["payout_status"] == "partial"),
        "paid_count": sum(1 for r in rows if r["payout_status"] == "paid"),
    }


def record_payment(deal_agent_id, amount, paid_at, note, user_id):
    """Add a payment against one deal_agents row. Amounts accumulate, so a
    partial payment followed by another just brings the balance down —
    nothing is overwritten."""
    row = query("SELECT * FROM deal_agents WHERE id = ?", (deal_agent_id,), one=True)
    if row is None:
        return None
    new_paid = max(0, (row["paid_amount"] or 0) + amount)
    status = _status(row["amount"] or 0, new_paid)
    execute(
        "UPDATE deal_agents SET paid_amount=?, payout_status=?, paid_at=?,"
        " payout_note=?, recorded_by=? WHERE id=?",
        (new_paid, status, paid_at, note or None, user_id, deal_agent_id))
    log(user_id, "Recorded a commission payout", "deal", row["deal_id"],
        f"{amount:,.2f} to user #{row['user_id']}")
    return status


REQUIRED_COLUMNS = ("deal ref", "agent email", "amount paid")


def import_payouts_file(path, user_id):
    """Bulk-apply payouts from an Excel sheet: Deal Ref, Agent Email, Amount
    Paid, and optionally Paid Date / Note. Matches each row to a deal_agents
    record by deal ref + agent email so a mistyped ref or email is reported
    back rather than silently creating something new — there is nothing to
    create here, only existing earnings to mark as paid."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
    except StopIteration:
        return {"error": "That file looks empty."}
    header = [str(c.value or "").strip().lower() for c in header_row]

    def col(*names):
        for name in names:
            if name in header:
                return header.index(name)
        return None

    i_ref = col("deal ref", "ref")
    i_email = col("agent email", "email")
    i_amt = col("amount paid", "amount", "paid amount")
    i_date = col("paid date", "date")
    i_note = col("note", "notes")
    if i_ref is None or i_email is None or i_amt is None:
        return {"error": "The file needs at least Deal Ref, Agent Email and "
                         "Amount Paid columns."}

    updated, not_found, skipped = 0, [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c in (None, "") for c in row):
            continue
        ref = str(row[i_ref] or "").strip()
        email = str(row[i_email] or "").strip().lower()
        try:
            amount = float(row[i_amt] or 0)
        except (TypeError, ValueError):
            amount = 0
        if not ref or not email or amount <= 0:
            skipped += 1
            continue

        da = query(
            "SELECT da.id FROM deal_agents da"
            " JOIN deals d ON d.id = da.deal_id"
            " JOIN users u ON u.id = da.user_id"
            " WHERE d.ref = ? AND lower(u.email) = ?", (ref, email), one=True)
        if da is None:
            not_found.append(f"{ref} / {email}")
            continue

        raw_date = row[i_date] if i_date is not None else None
        paid_at = (to_utc(str(raw_date)) if raw_date else None) or now()
        note = (str(row[i_note]).strip() if i_note is not None and row[i_note]
                else "Imported from Excel")
        record_payment(da["id"], amount, paid_at, note, user_id)
        updated += 1

    return {"updated": updated, "not_found": not_found, "skipped": skipped}
