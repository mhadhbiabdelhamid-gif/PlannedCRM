"""Reading partner availability lists.

Every agency sends a different shape: headers three rows down under a logo,
merged section rows, unit numbers buried in codes like ARPQ02-B00-F01-A101,
rents written as "9000 qrs", sizes as "113 sqm". This module works out the
layout, guesses which column is which, and pulls out clean values. Nothing is
saved until a person has looked at the preview and agreed.
"""
import itertools
import re

from openpyxl import load_workbook

# Fields the CRM can fill from a spreadsheet.
FIELDS = [
    ("unit_no", "Flat / unit number"),
    ("building_no", "Building"),
    ("floor_no", "Floor"),
    ("title", "Title or property name"),
    ("prop_type", "Property type"),
    ("bedrooms", "Bedrooms"),
    ("bathrooms", "Bathrooms"),
    ("size_sqm", "Size"),
    ("price", "Rent or price"),
    ("status", "Status"),
    ("area", "Location / district"),
    ("description", "Description or notes"),
    ("map_url", "Map link"),
    ("features", "Features, view, furnishing"),
    ("extras", "Extra rooms (office, maid's, balcony)"),
]

# Words that suggest a column holds a given field. Matched against the header
# text, lowercased, longest first so "unit size" beats "unit".
SYNONYMS = {
    "unit_no": ["unit number", "unit no", "apt no", "apartment no", "flat no",
                "unit #", "unit", "apartment number", "flat number", "door no"],
    "building_no": ["building name", "property name", "building no", "building",
                    "tower", "block", "compound", "project"],
    "floor_no": ["floor number", "floor no", "floor", "level", "storey", "story",
                 "flr"],
    "extras": ["extra rooms", "additional rooms", "extras", "maid room",
               "maid's room", "additional"],
    "title": ["property", "description of property", "name", "unit type name"],
    "prop_type": ["property type", "apartment type", "type of property", "type",
                  "category"],
    "bedrooms": ["no. bedroom", "no of bedroom", "number of bedrooms", "bedrooms",
                 "bedroom", "no. of beds", "beds", "bed", "layout", "bhk",
                 "configuration"],
    "bathrooms": ["bathrooms", "bathroom", "baths", "bath", "no. of baths"],
    "size_sqm": ["unit size", "size sqm", "built up area", "area sqm", "sqm",
                 "sq.m", "sq m", "size", "built-up", "bua"],
    "price": ["monthly rent", "rent per month", "asking price", "rate", "rent",
              "price", "amount", "monthly", "sale price"],
    "status": ["booking status", "availability", "status", "vacant", "available"],
    "area": ["location", "district", "area", "zone", "neighbourhood",
             "neighborhood", "address"],
    "description": ["description", "remarks", "notes", "comment", "details"],
    "map_url": ["map", "google map", "location link", "maps link", "pin"],
    "features": ["view", "furniture", "furnishing", "balcony", "features",
                 "amenities", "facing"],
}

# Header text that means the row is a header rather than data.
HEADER_HINTS = set()
for words in SYNONYMS.values():
    HEADER_HINTS.update(words)

NOISE = re.compile(r"[\s\u00a0]+")


def clean(value):
    if value is None:
        return ""
    return NOISE.sub(" ", str(value)).strip()


# ------------------------------------------------------------------ parsing
WORD_NUMBERS = {
    "studio": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
    "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
}


# Every way a partner might write a bedroom count. Longest first so "bedroom"
# is matched before "bed", and "bdrm" before "bd".
BED_WORDS = (r"bedrooms?|bed\s*rooms?|bedrms?|bdrms?|bdms?|bdrs?|"
             r"b\s*/\s*r|bhk|beds?|bds?|brs?|rooms?|غرف(?:ة|تين)?|غرفه")

# Extra rooms that are not bedrooms, so "1 bd + off" is one bedroom, not two.
EXTRAS = re.compile(
    r"\+?\s*\b(off(?:ice)?|study|maid[' ]?s?(?:\s*room)?|hall|majlis|store|"
    r"storage|driver[' ]?s?(?:\s*room)?|laundry|balcony|terrace|garden|"
    r"pantry|nanny)\b", re.IGNORECASE)

STUDIO_WORDS = re.compile(r"\bstudios?\b|\bstd\b|\bستوديو\b", re.IGNORECASE)

# Units and words that follow a number but have nothing to do with bedrooms.
# Without these, "50 sqm" was being read as fifty bedrooms.
NOT_BEDROOMS = re.compile(
    r"\b(sq\.?\s*m|sqm|sqft|sq\.?\s*ft|m2|m²|ft2|meters?|metres?|"
    r"qar?|qr|riyals?|aed|usd|k|kw|kwh|"
    r"floors?|flr|storey?s?|levels?|"
    r"parking|car\s*parks?|spaces?|years?|months?|days?|"
    r"units?|shops?|offices?)\b", re.IGNORECASE)

MAX_BEDROOMS = 15          # beyond this it is not a flat, it is a data error


def parse_bedrooms(value):
    """Read a bedroom count out of however the partner wrote it.

    Handles '1 bd+off', '2BHK+Maid', 'Studio', 'Two Bedrooms', '3 Beds; 5 Baths',
    '1-BR', '2 bdr', '3 R', '2 غرفة' and a plain number. Extra rooms — office,
    maid's, study — are recognised so they are not counted as bedrooms.

    Returns None when there is genuinely no count, never 0 as a guess, so the
    review screen can flag it rather than quietly inventing a studio.
    """
    text = clean(value)
    if not text:
        return None

    if STUDIO_WORDS.search(text):
        return 0

    # strip the extras first, so their words cannot be mistaken for a count
    stripped = EXTRAS.sub(" ", text)

    # a digit attached to a bedroom word: "2BHK", "1 bd", "3-BR", "4 bed rooms"
    m = re.search(rf"(\d+)\s*[-+/]?\s*(?:{BED_WORDS})\b", stripped, re.IGNORECASE)
    if m:
        return _sane(int(m.group(1)))

    # the word before the digit: "bedrooms: 3", "BR 2"
    m = re.search(rf"(?:{BED_WORDS})\s*[:=-]?\s*(\d+)", stripped, re.IGNORECASE)
    if m:
        return _sane(int(m.group(1)))

    # spelled out: "Two Bedrooms", "Three bedroom villa", "Two BR"
    for word, n in WORD_NUMBERS.items():
        if re.search(rf"\b{word}\b\s*[-]?\s*(?:{BED_WORDS})\b", stripped,
                     re.IGNORECASE):
            return n

    # anything measured in square metres, riyals, floors and so on is not a
    # bedroom count, however it is written
    if NOT_BEDROOMS.search(stripped):
        return None

    # a bare count where the whole cell is a number: "2", "0"
    m = re.match(r"^\s*(\d{1,2})\s*$", stripped)
    if m:
        return _sane(int(m.group(1)))

    # "2+1" — the first figure is the bedrooms, the second the living rooms
    m = re.match(r"^\s*(\d{1,2})\s*\+\s*\d{1,2}\s*$", stripped)
    if m:
        return _sane(int(m.group(1)))

    # a lone digit next to nothing else meaningful: "3 R", "4 rm"
    m = re.match(r"^\s*(\d{1,2})\s*(?:r|rm|rms|غ)?\s*$", stripped, re.IGNORECASE)
    if m:
        return _sane(int(m.group(1)))

    return None


def _sane(n):
    """A flat with forty bedrooms means the column was the wrong one."""
    return n if n is not None and 0 <= n <= MAX_BEDROOMS else None


def describe_extras(value):
    """The extra rooms mentioned, so they land in features instead of vanishing."""
    text = clean(value)
    if not text:
        return ""
    found = []
    for m in EXTRAS.finditer(text):
        word = m.group(1).strip().title()
        word = {"Off": "Office", "Maid": "Maid's room",
                "Maids": "Maid's room", "Driver": "Driver's room"}.get(word, word)
        if word not in found:
            found.append(word)
    return ", ".join(found)


BATH_WORDS = r"bathrooms?|bath\s*rooms?|bathrms?|baths?|washrooms?|wc|w\.c\.|ba\b|حمام(?:ات)?"


def parse_bathrooms(value):
    text = clean(value)
    if not text:
        return None
    m = re.search(rf"(\d+)\s*[-+/]?\s*(?:{BATH_WORDS})", text, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(rf"(?:{BATH_WORDS})\s*[:=-]?\s*(\d+)", text, re.IGNORECASE)
    if m:
        return int(m.group(1))
    for word, n in WORD_NUMBERS.items():
        if re.search(rf"\b{word}\b\s*[-]?\s*(?:{BATH_WORDS})", text, re.IGNORECASE):
            return n
    return None


def parse_number(value):
    """Pulls a figure out of '9000 qrs', 'QR 8,500', '113 sqm', 8500."""
    if isinstance(value, (int, float)):
        return float(value)
    text = clean(value)
    if not text:
        return None
    text = text.replace(",", "")
    m = re.search(r"(\d+(?:\.\d+)?)", text)
    return float(m.group(1)) if m else None


UNIT_CODE = re.compile(r"[A-Z]{2,}\d*-[A-Z0-9]+-[A-Z0-9]+-([A-Z]*\d+[A-Z]*)",
                       re.IGNORECASE)


def parse_unit(value):
    """Unit numbers arrive in several disguises.

    'ARPQ02-B00-F01-A101' -> A101      (the last segment is the door number)
    '311 / Balcony'       -> 311
    'Villa No. A-15 (inside)' -> A-15
    'Flat No. 48'         -> 48
    """
    text = clean(value)
    if not text:
        return ""

    m = UNIT_CODE.search(text)
    if m:
        return m.group(1).upper()

    # strip a leading label
    text = re.sub(r"^(villa|flat|apartment|apt|unit|shop|office)\s*(no\.?|#|number)?\s*",
                  "", text, flags=re.IGNORECASE).strip()
    # drop anything after a separator: "311 / Balcony", "1707 No Balcony"
    text = re.split(r"\s*[/|,]\s*|\s{2,}", text)[0].strip()
    text = re.sub(r"\((.*?)\)", "", text).strip()
    m = re.match(r"^([A-Za-z]?[-\s]?\d+[A-Za-z]?(?:-\d+)?)", text)
    if m:
        return m.group(1).replace(" ", "").upper()
    return text[:20]


FLOOR_WORDS = {
    "ground": "G", "gf": "G", "g": "G", "mezzanine": "M", "mezz": "M",
    "basement": "B", "penthouse": "PH", "roof": "R",
}

ORDINAL = re.compile(r"^\s*(\d{1,3})\s*(?:st|nd|rd|th)?\s*(?:floor|flr|level)?\s*$",
                     re.IGNORECASE)
FLOOR_IN_CODE = re.compile(r"-F(\d{1,3})-", re.IGNORECASE)


def parse_floor(value, unit_code=""):
    """'1st', 'Ground', 'G', 'B1', 'Floor 12', or hidden inside a unit code
    like ARPQ02-B00-F01-A101, where F01 is the first floor."""
    text = clean(value)
    if not text:
        m = FLOOR_IN_CODE.search(clean(unit_code))
        return str(int(m.group(1))) if m else ""

    low = text.lower().strip(" .-")
    if low in FLOOR_WORDS:
        return FLOOR_WORDS[low]
    for word, short in FLOOR_WORDS.items():
        if re.fullmatch(rf"{word}\s*floor", low):
            return short

    m = ORDINAL.match(text)
    if m:
        return str(int(m.group(1)))
    m = re.search(r"(?:floor|flr|level|storey|story)\s*[:.\-]?\s*(\d{1,3})",
                  text, re.IGNORECASE)
    if m:
        return str(int(m.group(1)))
    m = re.match(r"^\s*([BM]\s?\d{1,2}|PH\d?)\s*$", text, re.IGNORECASE)
    if m:
        return m.group(1).upper().replace(" ", "")
    return text[:12]


STATUS_WORDS = [
    (("vacant", "available", "ready", "free", "rfo"), "Available"),
    (("booked", "reserved", "on hold", "hold", "under offer"), "Reserved"),
    (("rented", "leased", "occupied", "let"), "Rented"),
    (("sold",), "Sold"),
]


def parse_status(value, default="Available"):
    text = clean(value).lower()
    if not text:
        return default
    for words, status in STATUS_WORDS:
        if any(w in text for w in words):
            return status
    return default


TYPE_WORDS = [
    (("villa", "compound", "townhouse"), "Villa"),
    (("office", "commercial office"), "Office"),
    (("retail", "shop", "showroom", "commercial", "warehouse"), "Commercial"),
    (("land", "plot"), "Land"),
    (("apartment", "flat", "studio", "bhk", "br", "penthouse", "residential"),
     "Apartment"),
]


def parse_type(*values, default="Apartment"):
    joined = " ".join(clean(v).lower() for v in values if clean(v))
    for words, kind in TYPE_WORDS:
        if any(w in joined for w in words):
            return kind
    return default


MAP_RE = re.compile(r"https?://\S*(?:google\.[a-z.]+/maps|maps\.app\.goo\.gl|"
                    r"goo\.gl/maps|maps\.google)\S*", re.IGNORECASE)


def find_map_link(*values):
    for v in values:
        text = clean(v)
        m = MAP_RE.search(text)
        if m:
            return m.group(0).rstrip(").,;")
    return ""


# -------------------------------------------------------- layout detection
def score_header_row(cells):
    """How much does this row look like a set of column headings?"""
    score = 0
    filled = 0
    for cell in cells:
        text = clean(cell).lower()
        if not text:
            continue
        filled += 1
        if len(text) > 45:                 # a sentence, not a heading
            score -= 2
            continue
        for hint in HEADER_HINTS:
            if hint in text or text in hint:
                score += 3
                break
        else:
            if not re.match(r"^[\d.,\s]+$", text):
                score += 0.5               # a short word is plausible
    if filled < 2:
        return -10
    return score


def detect_header_row(ws, limit=20, col_range=None):
    c_start, c_end = col_range or (1, min(ws.max_column, 25))
    best, best_score = 1, -99
    for r in range(1, min(ws.max_row, limit) + 1):
        cells = [ws.cell(row=r, column=c).value
                 for c in range(c_start, c_end + 1)]
        s = score_header_row(cells)
        if s > best_score:
            best, best_score = r, s
    return best


def find_column_blocks(ws, gap=2, scan_rows=60, max_col=60, max_block_width=25):
    """Partners sometimes paste two buildings' lists side by side on one
    sheet instead of stacking them, so the second table's columns start
    partway across the row rather than at column A. Treated as one table,
    that column offset gets read as extra columns of the first table and
    the second table's own header is read as a data row — its fields never
    get mapped, so most of it goes missing.

    Detects that by columns: a run of `gap` or more columns that are blank
    across every one of the first `scan_rows` rows is treated as a divider
    between separate tables (a single blank spacer column inside one table
    is normal and doesn't split anything). A normal single-table sheet
    always returns one block spanning the whole width, so this changes
    nothing for the common case.
    """
    ncols = min(ws.max_column, max_col)
    if ncols < 1:
        return [(1, 1)]
    last_row = min(ws.max_row, scan_rows) or 1
    empty = [all(not clean(ws.cell(row=r, column=c).value)
                 for r in range(1, last_row + 1))
             for c in range(1, ncols + 1)]

    blocks = []
    col = 1
    cur_start = None
    for is_empty, group in itertools.groupby(empty):
        length = sum(1 for _ in group)
        if is_empty:
            if length >= gap and cur_start is not None:
                blocks.append((cur_start, col - 1))
                cur_start = None
        elif cur_start is None:
            cur_start = col
        col += length
    if cur_start is not None:
        blocks.append((cur_start, ncols))
    if not blocks:
        blocks = [(1, ncols)]

    # A sliver with almost nothing in it is stray notes, not a second table.
    real = []
    for s, e in blocks:
        filled = sum(1 for r in range(1, last_row + 1) for c in range(s, e + 1)
                     if clean(ws.cell(row=r, column=c).value))
        if filled >= 3:
            real.append((s, min(e, s + max_block_width - 1)))
    return real or blocks


# Fields where several columns should be joined together rather than one
# winning. A partner may split an address or a description across columns.
JOINABLE = {"title", "description", "features", "address", "area",
            "building_no"}


def as_columns(value):
    """Mapping values may be a single column or several. Always return a list."""
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        return [int(v) for v in value if str(v).strip().isdigit() and int(v) > 0]
    return [int(value)] if str(value).strip().isdigit() and int(value) > 0 else []


def guess_mapping(headers):
    """headers: list of header strings. Returns {field: column index (1-based)}."""
    mapping = {}
    used = set()
    # longest synonyms first, so "unit size" wins over "unit"
    ordered = sorted(
        ((field, syn) for field, syns in SYNONYMS.items() for syn in syns),
        key=lambda pair: -len(pair[1]))

    for field, syn in ordered:
        if field in mapping:
            continue
        for i, head in enumerate(headers, start=1):
            if i in used:
                continue
            text = clean(head).lower()
            if not text:
                continue
            if text == syn or text.startswith(syn) or syn in text:
                mapping[field] = i
                used.add(i)
                break
    return mapping


UNIT_LABEL = re.compile(r"^\s*(villa|flat|apartment|apt|unit|shop|office|studio)\b",
                        re.IGNORECASE)


def infer_from_values(headers, rows, mapping):
    """Some columns have no heading at all — the Pearl list keeps bedroom
    descriptions ('Studio', '1br+Off') under a blank header. Judge those by
    what is in them instead."""
    used = set(mapping.values())
    for i in range(1, len(headers) + 1):
        if i in used:
            continue
        sample = [clean(r[i - 1]) for r in rows[:25] if i <= len(r)]
        sample = [v for v in sample if v]
        if len(sample) < 2:
            continue

        if "bedrooms" not in mapping:
            hits = sum(1 for v in sample
                       if re.search(r"studio|\d\s*(br|bhk|bed)", v, re.IGNORECASE))
            if hits >= max(2, len(sample) * 0.6):
                mapping["bedrooms"] = i
                used.add(i)
                continue

        if "status" not in mapping:
            words = {w for group, _ in STATUS_WORDS for w in group}
            hits = sum(1 for v in sample if v.lower() in words)
            if hits >= max(2, len(sample) * 0.7):
                mapping["status"] = i
                used.add(i)
    return mapping


STRONG = ("building", "tower", "compound", "block", "project")
WEAK = ("residence", "villa", "apartments", "gardens", "plaza")


def guess_context(ws, header_row, col_range=None):
    """Partners often name the building in a title row above the table rather
    than giving it a column. Rows nearest the table win, and a line saying
    'Building' beats a general banner."""
    c_start, c_end = col_range or (1, min(ws.max_column, 8))
    c_end = min(c_end, c_start + 7)
    best, best_score = "", 0
    for r in range(header_row - 1, 0, -1):
        for c in range(c_start, c_end + 1):
            text = clean(ws.cell(row=r, column=c).value)
            if not text or len(text) < 4 or len(text) > 60:
                continue
            if re.match(r"^[\d\s./-]+$", text):
                continue
            low = text.lower()
            if any(w in low for w in ("list", "availability", "rates", "including",
                                      "coming soon", "updated", "note")):
                continue                      # a banner, not a building
            score = 1
            if any(w in low for w in STRONG):
                score = 4
            elif any(w in low for w in WEAK):
                score = 2
            score += (header_row - r) * 0.01   # nearer the table is better
            if score > best_score:
                best, best_score = text, score
    return best


def read_sheet(ws, header_row=None, col_range=None):
    """Everything we know about one sheet, before any decisions are made."""
    c_start, c_end = col_range or (1, min(ws.max_column, 25))
    header_row = header_row or detect_header_row(ws, col_range=(c_start, c_end))
    headers = [clean(ws.cell(row=header_row, column=c).value)
               for c in range(c_start, c_end + 1)]
    rows, links = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        values, targets = [], []
        for c in range(c_start, c_end + 1):
            cell = ws.cell(row=r, column=c)
            values.append(cell.value)
            # A partner often writes "Compound - Google Maps" with the real
            # address hidden behind it, so read the link as well as the text.
            if cell.hyperlink is not None and cell.hyperlink.target:
                targets.append(cell.hyperlink.target)
        if any(clean(v) for v in values):
            rows.append(values)
            links.append(targets)
    mapping = infer_from_values(headers, rows, guess_mapping(headers))
    return {"header_row": header_row, "headers": headers, "rows": rows,
            "links": links, "mapping": mapping,
            "context": guess_context(ws, header_row, col_range=(c_start, c_end)),
            "col_range": (c_start, c_end)}


def read_sheet_blocks(ws, header_row=None):
    """Like read_sheet, but first checks whether the sheet actually holds more
    than one table placed side by side (see find_column_blocks) and reads
    each as its own table, with its own header row and column mapping, so a
    second table's data isn't dropped or blended into the first table's
    rows. Returns a list of infos — almost every sheet yields exactly one,
    identical to what read_sheet(ws) would have returned on its own.

    A header row chosen by hand only ever applies to the first table; later
    tables keep using their own auto-detected header, since there is no way
    for a single override to mean two different rows.
    """
    blocks = find_column_blocks(ws)
    if len(blocks) <= 1:
        return [read_sheet(ws, header_row=header_row)]
    infos = [read_sheet(ws, header_row=header_row, col_range=blocks[0])]
    infos += [read_sheet(ws, col_range=b) for b in blocks[1:]]
    return infos


# ------------------------------------------------------------- extraction
def extract(sheet, mapping, defaults, fill_down=True, fill_numbers=False):
    """Turn raw rows into listings ready for review.

    fill_down copies a blank building, location or description from the row
    above — the usual shape when a partner lists several units under one
    heading. fill_numbers does the same for rent, bedrooms and size, which is
    right for hierarchical lists and wrong for flat ones, so it is off by
    default.
    """
    out = []
    carried = {}
    seen_in_file = set()
    # Structural facts repeat down a list and are safe to carry. A description
    # belongs to one unit, so it only carries when the fuller option is chosen.
    inherit = ["building_no", "area", "map_url", "prop_type"]
    if fill_numbers:
        inherit += ["price", "bedrooms", "bathrooms", "size_sqm", "description"]

    def cell(values, field):
        """One field may draw on more than one column.

        Text fields join what they find, so a description split across three
        columns arrives whole. Everything else takes the first column that
        actually holds something, so an empty column does not mask a later one.
        """
        columns = as_columns(mapping.get(field))
        picked = []
        for idx in columns:
            if idx > len(values):
                continue
            value = values[idx - 1]
            if clean(value):
                picked.append(value)
        if not picked:
            return None
        if field in JOINABLE and len(picked) > 1:
            seen, parts = set(), []
            for v in picked:
                text = clean(v)
                if text.lower() not in seen:
                    seen.add(text.lower())
                    parts.append(text)
            return " · ".join(parts)
        return picked[0]

    all_links = sheet.get("links") or [[] for _ in sheet["rows"]]
    for index, values in enumerate(sheet["rows"]):
        row_links = all_links[index] if index < len(all_links) else []
        raw = {f: cell(values, f) for f, _ in FIELDS}
        # remember what was actually in the row before anything is carried down
        own_building = clean(raw.get("building_no"))
        own_title = clean(raw.get("title"))
        own_unit = parse_unit(raw.get("unit_no"))
        own_price = parse_number(raw.get("price"))
        own_size = parse_number(raw.get("size_sqm"))

        if fill_down:
            for field in inherit:
                if clean(raw.get(field)):
                    carried[field] = raw[field]
                elif field in carried:
                    raw[field] = carried[field]

        unit = parse_unit(raw.get("unit_no"))
        title_text = clean(raw.get("title"))
        price = parse_number(raw.get("price"))
        beds = parse_bedrooms(raw.get("bedrooms"))
        if beds is None:
            beds = parse_bedrooms(raw.get("prop_type"))
        if beds is None:
            beds = parse_bedrooms(title_text)
        size = parse_number(raw.get("size_sqm"))

        # A row with no unit, no price and no size is a section heading or a
        # footnote — unless its text names a unit ("Villa No. A-16"), which is
        # how some lists write a unit whose rent matches the one above.
        looks_like_unit = bool(UNIT_LABEL.match(own_title))
        if (not own_unit and own_price is None and own_size is None
                and not looks_like_unit):
            # a heading names the building for the rows beneath it
            if own_title and not own_building:
                carried["building_no"] = own_title
                raw["building_no"] = own_title
            heading_link = find_map_link(*row_links, *values)
            if heading_link:
                carried["map_url"] = heading_link
            continue

        building = clean(raw.get("building_no")) or defaults.get("building_no", "")
        # some sheets put the unit label in the title column instead
        if not unit and title_text:
            unit = parse_unit(title_text)

        prop_type = parse_type(raw.get("prop_type"), raw.get("bedrooms"),
                               title_text, building,
                               default=defaults.get("prop_type", "Apartment"))

        # "1 bd + office" means one bedroom and an office; keep the office
        # Extra rooms are their own field now, so "1 bd + office" records the
        # office as a room rather than burying it in the features text.
        found_extras = [clean(raw.get("extras"))]
        for source in (raw.get("bedrooms"), title_text, raw.get("prop_type")):
            more = describe_extras(source)
            if more:
                found_extras.append(more)
        seen, parts = set(), []
        for chunk in found_extras:
            for piece in re.split(r"[,;·]", chunk or ""):
                piece = piece.strip()
                if piece and piece.lower() not in seen:
                    seen.add(piece.lower())
                    parts.append(piece)
        extras_text = ", ".join(parts)
        features = clean(raw.get("features"))

        listing = {
            "unit_no": unit,
            "building_no": building,
            "floor_no": parse_floor(raw.get("floor_no"),
                                    clean(raw.get("unit_no")) or title_text),
            "prop_type": prop_type,
            "bedrooms": beds,
            "bathrooms": parse_bathrooms(raw.get("bathrooms")) or
                         parse_bathrooms(raw.get("bedrooms")),
            "size_sqm": size,
            "price": price or 0,
            "status": parse_status(raw.get("status"),
                                   defaults.get("status", "Available")),
            "area": clean(raw.get("area")) or defaults.get("area", ""),
            "description": clean(raw.get("description")),
            "features": features,
            "extras": extras_text,
            # any column can hold the pin, and it may be a hyperlink rather
            # than visible text
            "map_url": find_map_link(raw.get("map_url"), *row_links, *values),
            "listing_type": defaults.get("listing_type", "Rent"),
        }

        bits = [b for b in (building, f"unit {unit}" if unit else "") if b]
        listing["title"] = (title_text if title_text and not unit
                            else " · ".join(bits) or title_text or "Untitled listing")
        if beds is not None and listing["title"] and unit:
            label = "Studio" if beds == 0 else f"{beds}-bed"
            listing["title"] = f"{label} · {' · '.join(bits)}"

        # Flag anything a person should look at before it is saved.
        issues = []
        if not unit and not building:
            issues.append("no building or flat number")
        if not listing["price"]:
            issues.append("no price")
        if listing["bedrooms"] is None and listing["prop_type"] == "Apartment":
            issues.append("no bedroom count")
        if clean(raw.get("map_url")) and not listing["map_url"]:
            issues.append("map link not recognised")
        listing["issues"] = issues

        key = (building.strip().lower(), unit.strip().lower())
        if key in seen_in_file and any(key):
            listing["issues"] = issues + ["appears twice in this file"]
        seen_in_file.add(key)

        out.append(listing)
    return out


def open_workbook(path):
    return load_workbook(path, data_only=True)
