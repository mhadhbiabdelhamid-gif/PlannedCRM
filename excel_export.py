"""Branded Excel exports.

Produces the workbook agreed in the template draft: a company header block
with the logo on every sheet, a printed footer, and the property listings
split into ours and everyone else's. Company details come from Settings, so
changing the address there changes every future export.

Every sheet is sized to print cleanly on A4 (fit to page width, repeating
header rows) and, when the person exporting is using the CRM in Arabic, the
whole workbook — headers, titles, statuses, footer — comes out in Arabic
with the sheet mirrored right-to-left, not just the numbers.
"""
import io
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from db import get_setting, local_now, to_local
from i18n import is_rtl, t

try:
    from openpyxl.drawing.image import Image as XLImage
    HAS_LOGO_SUPPORT = True
except ImportError:
    # openpyxl needs Pillow to read/embed an image. Pillow is optional
    # elsewhere in this app (see README), so exports must keep working
    # without it — they just come out without a logo.
    HAS_LOGO_SUPPORT = False

# ------------------------------------------------------------------ brand
GOLD = "C8A24A"
GOLD_PALE = "F6EFDD"
INK = "0B0B0D"
PAPER = "F4F1EA"
LINE = "DBD7CE"
MUTED = "4E4B44"
LINK = "0563C1"
GOOD_PALE = "E8F3E7"
WARN_PALE = "FBF0DC"
BAD_PALE = "FBEAEA"
FONT = "Arial"

thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADER_ROW = 8
FIRST_DATA = 9

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_EXTS = ("png", "jpg", "jpeg", "webp", "gif")  # raster only — SVG can't
                                                    # be embedded via Pillow


def _find_logo_path():
    """Same lookup as app.py's find_logo(), independent of it — importing
    app.py here would be circular (app -> views_* -> excel_export)."""
    folder = os.path.join(BASE_DIR, "static", "img")
    if not os.path.isdir(folder):
        return None
    for entry in sorted(os.listdir(folder)):
        stem, _, ext = entry.rpartition(".")
        if stem.lower() == "logo" and ext.lower() in LOGO_EXTS:
            return os.path.join(folder, entry)
    return None


def company():
    return {
        "name": get_setting("company_name", "Planned Real Estate"),
        "tagline": get_setting("tagline", "Property Consultants · Doha, Qatar"),
        "address": get_setting("address", ""),
        "po_box": get_setting("po_box", ""),
        "phone": get_setting("phone", ""),
        "phone2": get_setting("phone2", ""),
        "email": get_setting("email", ""),
        "cr": get_setting("cr_number", ""),
        "currency": get_setting("currency", "QAR"),
    }


def _join(parts, sep="   ·   "):
    return sep.join(p for p in parts if p)


def _side(horizontal, rtl):
    """Flip a plain left/right alignment for an RTL sheet. Center is left
    alone; numeric columns that were deliberately set to 'right' for LTR
    reading stay 'right' — only the text columns that default to 'left'
    (the reading-start side) need to flip to the RTL reading-start side."""
    if rtl and horizontal == "left":
        return "right"
    return horizontal


def _date_fmt(rtl):
    # English month abbreviations ("19 Aug 2026") read oddly inside an
    # otherwise Arabic document, so Arabic exports get a numeric date.
    return "dd/mm/yyyy" if rtl else "dd mmm yyyy"


# strftime's %B is locale-bound (usually English no matter what language the
# CRM is set to), so the "Exported ..." stamp is built from this fixed
# English list and routed through t() instead — same trick as reports.py's
# period_label(), so an Arabic export gets an Arabic month name too.
_MONTHS_FULL = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]


def brand_header(ws, co, title, subtitle, ncols, exported_by, rtl=False):
    last = get_column_letter(ncols)
    logo_path = _find_logo_path() if HAS_LOGO_SUPPORT else None

    # The logo gets its own gutter in columns 1-2 of the top band only, so
    # it never sits under the company name text. Every other row still
    # spans the full sheet width as before.
    name_col = min(3, ncols) if logo_path else 1
    name_start = get_column_letter(name_col)

    if logo_path:
        for row in (1, 2):
            for col in range(1, name_col):
                c = ws.cell(row=row, column=col)
                c.fill = PatternFill("solid", fgColor=INK)
        try:
            img = XLImage(logo_path)
            target_h = 42
            scale = target_h / float(img.height)
            img.height = target_h
            img.width = int(img.width * scale)
            ws.add_image(img, "A1")
        except Exception:
            pass  # a corrupt/unreadable logo file must not break the export

    ws.merge_cells(f"{name_start}1:{last}1")
    c = ws[f"{name_start}1"]
    c.value = co["name"]
    c.font = Font(name=FONT, size=20, bold=True, color=GOLD)
    c.fill = PatternFill("solid", fgColor=INK)
    c.alignment = Alignment(horizontal=_side("left", rtl), vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f"{name_start}2:{last}2")
    c = ws[f"{name_start}2"]
    c.value = co["tagline"]
    c.font = Font(name=FONT, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=INK)
    c.alignment = Alignment(horizontal=_side("left", rtl), vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    lines = [
        _join([co["address"], co["po_box"]]),
        _join([f"{t('Tel')} {co['phone']}" if co["phone"] else "", co["phone2"],
               co["email"], f"{t('CR')} {co['cr']}" if co["cr"] else ""]),
    ]
    for row, text in zip((3, 4), lines):
        ws.merge_cells(f"A{row}:{last}{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = Font(name=FONT, size=8.5, color=MUTED)
        c.fill = PatternFill("solid", fgColor=GOLD_PALE)
        c.alignment = Alignment(horizontal=_side("left", rtl), vertical="center", indent=1)
        ws.row_dimensions[row].height = 15

    ws.merge_cells(f"A5:{last}5")
    c = ws["A5"]
    c.value = title
    c.font = Font(name=FONT, size=13, bold=True, color=INK)
    c.alignment = Alignment(horizontal=_side("left", rtl), vertical="center", indent=1)
    ws.row_dimensions[5].height = 24

    ws.merge_cells(f"A6:{last}6")
    c = ws["A6"]
    now_local = local_now()
    stamp = (f"{now_local.day:02d} {t(_MONTHS_FULL[now_local.month - 1])} "
            f"{now_local.year} {now_local.strftime('%H:%M')}")
    exported_line = f"{t('Exported')} {stamp} ({t('Doha time')}) {t('by')} {exported_by}"
    c.value = f"{subtitle}   |   {exported_line}" if subtitle else exported_line
    c.font = Font(name=FONT, size=8.5, italic=True, color=MUTED)
    c.alignment = Alignment(horizontal=_side("left", rtl), vertical="center", indent=1)
    ws.row_dimensions[6].height = 14

    ws.row_dimensions[7].height = 6


def table_header(ws, headers):
    for i, text in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=text)
        c.font = Font(name=FONT, size=9.5, bold=True, color=INK)
        c.fill = PatternFill("solid", fgColor=GOLD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 30


def write_rows(ws, rows, formats, rtl=False):
    for r, record in enumerate(rows, start=FIRST_DATA):
        banded = (r - FIRST_DATA) % 2 == 1
        for i, value in enumerate(record, start=1):
            c = ws.cell(row=r, column=i, value=value)
            c.font = Font(name=FONT, size=9.5, color="000000")
            c.border = BORDER
            c.alignment = Alignment(horizontal=_side(formats[i - 1][1], rtl),
                                    vertical="center")
            if formats[i - 1][0]:
                c.number_format = formats[i - 1][0]
            if banded:
                c.fill = PatternFill("solid", fgColor=PAPER)
    return FIRST_DATA + len(rows) - 1


def map_links(ws, col, first, last):
    """A short clickable label instead of a wall of URL.

    HYPERLINK() keeps the address readable in the formula bar and survives being
    copied elsewhere, which a plain hyperlink does not.
    """
    for r in range(first, last + 1):
        c = ws.cell(row=r, column=col)
        url = c.value
        if url:
            safe = str(url).replace('"', "%22")
            c.value = f'=HYPERLINK("{safe}","{t("Open map")}")'
            c.font = Font(name=FONT, size=9.5, color=LINK, underline="single")
        else:
            c.value = "—"
            c.font = Font(name=FONT, size=9.5, color="9B978E")
        c.alignment = Alignment(horizontal="center", vertical="center")


def totals_row(ws, row, ncols, label, formulas, rtl=False):
    ws.cell(row=row, column=1, value=label)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=9.5, bold=True, color=INK)
        cell.fill = PatternFill("solid", fgColor=GOLD_PALE)
        cell.border = Border(top=Side(style="medium", color=GOLD),
                             bottom=Side(style="medium", color=GOLD),
                             left=thin, right=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).alignment = Alignment(
        horizontal=_side("left", rtl), vertical="center", indent=1)
    for col, (formula, fmt) in formulas.items():
        cell = ws.cell(row=row, column=col, value=formula)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 20


def status_tint(ws, col, first, last, color_for):
    """Light background tint on one column by its own value — e.g. a payout
    status column, so unpaid/partial/paid stand out at a glance without a
    wall of colour across the whole row."""
    for r in range(first, last + 1):
        cell = ws.cell(row=r, column=col)
        color = color_for(cell.value)
        if color:
            cell.fill = PatternFill("solid", fgColor=color)


def finish(ws, co, ncols, widths, last_data, rtl=False):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=FIRST_DATA, column=1)
    if last_data >= FIRST_DATA:
        ws.auto_filter.ref = (f"A{HEADER_ROW}:"
                              f"{get_column_letter(ncols)}{last_data}")

    # A handful of narrow columns (a metrics summary, for instance) prints
    # far better as a single portrait A4 page than stretched landscape;
    # anything wider than that needs the extra room landscape gives it.
    ws.page_setup.orientation = "portrait" if ncols <= 4 else "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"1:{HEADER_ROW}"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.6)
    ws.print_options.horizontalCentered = True

    ws.sheet_view.rightToLeft = rtl

    ws.oddFooter.left.text = f"{co['name']} — {t('Confidential')}"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = "808080"
    ws.oddFooter.center.text = f"{t('Page')} &P {t('of')} &N"
    ws.oddFooter.center.size = 8
    ws.oddFooter.center.color = "808080"
    ws.oddFooter.right.text = "&F  ·  &D"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = "808080"

    ws.sheet_view.showGridLines = False


def empty_note(ws, ncols, text):
    ws.merge_cells(start_row=FIRST_DATA, start_column=1,
                   end_row=FIRST_DATA, end_column=ncols)
    c = ws.cell(row=FIRST_DATA, column=1, value=text)
    c.font = Font(name=FONT, size=10, italic=True, color=MUTED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[FIRST_DATA].height = 30


def _date(value):
    return to_local(value)


# ===================================================================== sheets
def property_sheet(wb, co, name, title, subtitle, records, exported_by,
                   index=None, rtl=False):
    headers = [t("Ref"), t("Title"), t("Building"), t("Floor"), t("Flat"),
               t("Location"), t("Map"), t("Type"), t("Sale / Rent"), t("Status"),
               f"{t('Price')} ({co['currency']})", t("Size (m²)"), t("Beds"),
               t("Baths"), t("Extra rooms"), t("Owner"), t("Agent"), t("Listed")]
    formats = [
        (None, "left"), (None, "left"), (None, "left"), ("@", "center"),
        ("@", "center"), (None, "left"), (None, "center"), (None, "center"),
        (None, "center"), (None, "center"), ("#,##0", "right"),
        ("#,##0.#", "right"), ("0", "center"), ("0", "center"), (None, "left"),
        (None, "left"), (None, "left"), (_date_fmt(rtl), "center"),
    ]
    widths = [11, 32, 16, 7, 8, 13, 10, 11, 10, 10, 14, 9, 6, 6, 22, 20, 16, 12]
    MAP_COL = 7

    ws = wb.create_sheet(t(name)) if index is None else wb.create_sheet(t(name), index)
    rows = [(r["ref"], r["title"], r["building_no"], r["floor_no"], r["unit_no"],
             r["area"], r["map_url"], t(r["prop_type"]), t(r["listing_type"]),
             t(r["status"]), r["price"], r["size_sqm"], r["bedrooms"],
             r["bathrooms"], r["extras"], r["owner_name"], r["agent_name"],
             _date(r["created_at"])) for r in records]

    brand_header(ws, co, title, subtitle.format(n=len(rows)), len(headers),
                exported_by, rtl)
    table_header(ws, headers)

    if rows:
        end = write_rows(ws, rows, formats, rtl)
        map_links(ws, MAP_COL, FIRST_DATA, end)
        n_label = f"{len(rows)} {t('listing') if len(rows) == 1 else t('listings')}"
        totals_row(ws, end + 1, len(headers), n_label, {
            11: (f"=SUM(K{FIRST_DATA}:K{end})", "#,##0"),
            12: (f"=SUM(L{FIRST_DATA}:L{end})", "#,##0.#"),
        }, rtl)
    else:
        end = FIRST_DATA - 1
        empty_note(ws, len(headers), t("No listings in this category yet."))
    finish(ws, co, len(headers), widths, end, rtl)


def leads_sheet(wb, co, records, exported_by, rtl=False):
    headers = [t("Ref"), t("Name"), t("Phone"), t("Email"), t("Source"), t("Stage"),
               f"{t('Budget')} ({co['currency']})", t("Property of interest"),
               t("Agent"), t("First contact"), t("Last update")]
    formats = [
        (None, "left"), (None, "left"), ("@", "left"), (None, "left"),
        (None, "center"), (None, "center"), ("#,##0", "right"), (None, "left"),
        (None, "left"), (_date_fmt(rtl), "center"), (_date_fmt(rtl), "center"),
    ]
    widths = [11, 22, 16, 26, 15, 12, 15, 30, 17, 13, 13]

    ws = wb.create_sheet(t("Leads"))
    rows = [(r["ref"], r["full_name"], r["phone"], r["email"], t(r["source"]),
             t(r["status"]), r["budget"], r["prop_title"], r["agent_name"],
             _date(r["created_at"]), _date(r["updated_at"])) for r in records]

    brand_header(ws, co, t("Leads Pipeline"),
                 f"{t('All leads')} · {t('every source')} · {t('every stage')} · "
                 f"{len(rows)} {t('records')}",
                 len(headers), exported_by, rtl)
    table_header(ws, headers)
    if rows:
        end = write_rows(ws, rows, formats, rtl)
        totals_row(ws, end + 1, len(headers), f"{len(rows)} {t('leads')}",
                   {7: (f"=SUM(G{FIRST_DATA}:G{end})", "#,##0")}, rtl)
    else:
        end = FIRST_DATA - 1
        empty_note(ws, len(headers), t("No leads captured yet."))
    finish(ws, co, len(headers), widths, end, rtl)


def deals_sheet(wb, co, records, exported_by, rtl=False):
    headers = [t("Ref"), t("Property"), t("Client"), t("Agent"), t("Type"),
               f"{t('Deal value')} ({co['currency']})", t("Comm. %"),
               f"{t('Commission')} ({co['currency']})", t("Status"), t("Closed on")]
    formats = [
        (None, "left"), (None, "left"), (None, "left"), (None, "left"),
        (None, "center"), ("#,##0", "right"), ("0.00%", "center"),
        ("#,##0.00", "right"), (None, "center"), (_date_fmt(rtl), "center"),
    ]
    widths = [11, 34, 21, 17, 10, 17, 10, 17, 12, 13]

    ws = wb.create_sheet(t("Deals"))
    rows = [(r["ref"], r["prop_title"], r["lead_name"], r["agent_name"],
             t(r["deal_type"]), r["value"], (r["commission_pct"] or 0) / 100.0,
             r["commission_amt"], t(r["status"]),
             _date(r["closed_at"] or r["created_at"])) for r in records]

    brand_header(ws, co, t("Deals and Commission"),
                 f"{t('All deals')} · {t('every status')} · {len(rows)} {t('records')}",
                 len(headers), exported_by, rtl)
    table_header(ws, headers)
    if rows:
        end = write_rows(ws, rows, formats, rtl)
        totals_row(ws, end + 1, len(headers), f"{len(rows)} {t('deals')}", {
            6: (f"=SUM(F{FIRST_DATA}:F{end})", "#,##0"),
            8: (f"=SUM(H{FIRST_DATA}:H{end})", "#,##0.00"),
        }, rtl)
    else:
        end = FIRST_DATA - 1
        empty_note(ws, len(headers), t("No deals recorded yet."))
    finish(ws, co, len(headers), widths, end, rtl)


ROLE_WORDS = {"admin": "Admin", "manager": "Manager", "accountant": "Accountant",
             "agent": "Agent"}


def agent_summary_sheet(wb, co, agent, period, deals, tasks, work, exported_by,
                        rtl=False):
    """Key figures for one agent's period, as a compact metric : value table —
    the same numbers the on-screen report shows, in a form that pastes
    cleanly into another spreadsheet or email."""
    headers = [t("Metric"), t("Value")]
    formats = [(None, "left"), (None, "right")]
    widths = [36, 20]

    rows = [
        (t("Deals closed"), deals["count"]),
        (f"{t('Deal value')} ({co['currency']})", deals["value"]),
        (f"{t('Commission earned')} ({co['currency']})", deals["commission"]),
        (f"{t('Commission collected')} ({co['currency']})", deals["collected"]),
        (t("Deals opened in period"), deals["opened"]),
        (t("Follow-ups due"), tasks["due_followups"]),
        (t("Follow-ups handled"), tasks["handled_followups"]),
        (t("Viewings scheduled"), tasks["viewings_scheduled"]),
        (t("Viewings completed"), tasks["viewings_done"]),
        (t("New leads"), work["new_leads"]),
        (t("Contacts logged"), work["contacts_logged"]),
        (t("New listings added"), work["new_listings"]),
        (t("Leads won"), work["won"]),
        (t("Leads lost"), work["lost"]),
    ]
    if "overdue_now" in tasks:
        rows.append((t("Overdue follow-ups (as of now)"), tasks["overdue_now"]))

    ws = wb.create_sheet(t("Summary"))
    role_word = t(ROLE_WORDS.get(agent["role"], "Agent"))
    brand_header(ws, co, f"{t('Agent Report')} — {agent['name']}",
                 f"{period['label']} · {agent['job_title'] or role_word}",
                 len(headers), exported_by, rtl)
    table_header(ws, headers)
    end = write_rows(ws, rows, formats, rtl)
    finish(ws, co, len(headers), widths, end, rtl)


def agent_deals_sheet(wb, co, agent, period, records, exported_by, rtl=False):
    headers = [t("Ref"), t("Property"), t("Client"), t("Type"),
               f"{t('Deal value')} ({co['currency']})", t("Comm. %"),
               f"{t('Commission')} ({co['currency']})", t("Status"), t("Closed on")]
    formats = [
        (None, "left"), (None, "left"), (None, "left"),
        (None, "center"), ("#,##0", "right"), ("0.00%", "center"),
        ("#,##0.00", "right"), (None, "center"), (_date_fmt(rtl), "center"),
    ]
    widths = [11, 34, 21, 10, 17, 10, 17, 12, 13]

    ws = wb.create_sheet(t("Deals"))
    rows = [(r["ref"], r["prop_title"], r["lead_name"], t(r["deal_type"]), r["value"],
             (r["commission_pct"] or 0) / 100.0, r["commission_amt"], t(r["status"]),
             _date(r["closed_at"] or r["created_at"])) for r in records]

    brand_header(ws, co, t("Deals closed in period"),
                 f"{agent['name']} · {period['label']} · {len(rows)} {t('records')}",
                 len(headers), exported_by, rtl)
    table_header(ws, headers)
    if rows:
        end = write_rows(ws, rows, formats, rtl)
        totals_row(ws, end + 1, len(headers), f"{len(rows)} {t('deals')}", {
            5: (f"=SUM(E{FIRST_DATA}:E{end})", "#,##0"),
            7: (f"=SUM(G{FIRST_DATA}:G{end})", "#,##0.00"),
        }, rtl)
    else:
        end = FIRST_DATA - 1
        empty_note(ws, len(headers), t("No deals closed in this period."))
    finish(ws, co, len(headers), widths, end, rtl)


def agent_report_workbook(report, agent, exported_by):
    """One agent's tasks/work/deals report (see reports.py) as a two-sheet
    workbook: a Summary of every figure, and the Deals that made up the
    period's closed total."""
    co = company()
    rtl = is_rtl()
    wb = Workbook()
    wb.remove(wb.active)

    period = report["period"]
    agent_summary_sheet(wb, co, agent, period, report["deals"], report["tasks"],
                        report["work"], exported_by, rtl)
    agent_deals_sheet(wb, co, agent, period, report["deals"]["rows"], exported_by, rtl)

    wb.properties.title = f"{co['name']} — {agent['name']} {t('report')}"
    wb.properties.creator = co["name"]
    wb.properties.created = datetime.now()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


PAYOUT_STATUS_WORDS = {"unpaid": "Unpaid", "partial": "Partially paid", "paid": "Paid"}
PAYOUT_TINT = {"Unpaid": BAD_PALE, "Partially paid": WARN_PALE, "Paid": GOOD_PALE}


def payouts_workbook(rows, totals, exported_by):
    """Commission payouts: what each agent earned on a live deal versus what
    they've actually been paid, for the accountant to reconcile."""
    co = company()
    rtl = is_rtl()
    headers = [t("Deal"), t("Property"), t("Agent"), t("Deal status"),
               f"{t('Earned')} ({co['currency']})", f"{t('Paid')} ({co['currency']})",
               t("Payout status"), t("Paid on"), t("Note")]
    formats = [
        (None, "left"), (None, "left"), (None, "left"), (None, "center"),
        ("#,##0.00", "right"), ("#,##0.00", "right"), (None, "center"),
        (_date_fmt(rtl), "center"), (None, "left"),
    ]
    widths = [11, 30, 20, 12, 16, 16, 14, 13, 26]

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(t("Payouts"))
    data = [(r["ref"], r["prop_title"], r["agent_name"], t(r["deal_status"]),
             r["amount"], r["paid_amount"],
             t(PAYOUT_STATUS_WORDS.get(r["payout_status"], "Unpaid")),
             _date(r["paid_at"]), r["payout_note"]) for r in rows]

    brand_header(ws, co, t("Commission Payouts"),
                 f"{t('Earned')} {co['currency']} {totals['earned']:,.0f} · "
                 f"{t('Paid')} {co['currency']} {totals['paid']:,.0f} · "
                 f"{t('Outstanding')} {co['currency']} {totals['outstanding']:,.0f} · "
                 f"{len(data)} {t('records')}",
                 len(headers), exported_by, rtl)
    table_header(ws, headers)
    if data:
        end = write_rows(ws, data, formats, rtl)
        status_tint(ws, 7, FIRST_DATA, end, lambda v: PAYOUT_TINT.get(v))
        totals_row(ws, end + 1, len(headers), f"{len(data)} {t('payouts')}", {
            5: (f"=SUM(E{FIRST_DATA}:E{end})", "#,##0.00"),
            6: (f"=SUM(F{FIRST_DATA}:F{end})", "#,##0.00"),
        }, rtl)
    else:
        end = FIRST_DATA - 1
        empty_note(ws, len(headers), t("No commission payouts recorded yet."))
    finish(ws, co, len(headers), widths, end, rtl)

    wb.properties.title = f"{co['name']} — {t('Commission payouts')}"
    wb.properties.creator = co["name"]
    wb.properties.created = datetime.now()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_workbook(properties, leads, deals, exported_by):
    """The whole workbook, in memory, ready to send to the browser."""
    co = company()
    rtl = is_rtl()
    wb = Workbook()
    wb.remove(wb.active)

    ours = [p for p in properties if p["is_own"]]
    theirs = [p for p in properties if not p["is_own"]]

    property_sheet(wb, co, "Our Properties", t("Our Properties"),
                   f"{t('Owned and managed by')} {co['name']} · {{n}} {t('listings')}",
                   ours, exported_by, rtl=rtl)
    property_sheet(wb, co, "Other Properties", t("Other Properties"),
                   f"{t('Third-party owners')} · {{n}} {t('listings')}",
                   theirs, exported_by, rtl=rtl)
    leads_sheet(wb, co, leads, exported_by, rtl)
    deals_sheet(wb, co, deals, exported_by, rtl)

    wb.properties.title = f"{co['name']} — {t('CRM export')}"
    wb.properties.creator = co["name"]
    wb.properties.created = datetime.now()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
